import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import pandas as pd
from collections import Counter
import os
from io import BytesIO

# SharePoint ETL functions import
from tools.sharepoint import load_env_vars, get_access_token, get_site_id, download_file

def create_assumption_summary_excel():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Assumption_Summary"
    
    data_structure = [
        ("Assumption Summary", "Input", "Accepted Format"),
        ("Project Name", "", ""),
        ("Data Tape with financials as of date:", "", "dd/mm/yyyy"),
        ("Valuation Date", "", "dd/mm/yyyy"),
        ("Output conclusions display currency", "", "E.g. EUR"),
        ("Data Tape currency - override", "", "E.g. EUR"),
        ("Detailed Guarantees Data Tape?", "", ""),
        ("Guarantee Currency - override", "", "E.g. EUR"),
        ("Global Tax Flag (Performing Loans only)", "", ""),
        ("Global Tax (Performing Loans only)", "", "%"),
        ("CoR Spread on Non-Collateralised loans", "", "%"),
        ("Assumed Maturities: Current Account", "", "Number of Months"),
        ("Assumed Maturities: Overdraft", "", "Number of Months"),
        ("Assumed Maturities: Credit Card", "", "Number of Months"),
        ("Type of Credit Card Curve Repayment to be used", "", ""),
        ("Assumption: % of the initial debt to be repaid - minimum Credit Card Monthly Payment", "", "%"),
        ("Cost of Risk Sensitivity Variance", "", "%"),
        ("Discount Rate Sensitivity Variance", "", "%"),
        ("Recovery Rate Sensitivity Variance", "", "%"),
        ("Sensitivity Table: Cost of Risk Sensitivity Range", "", "%"),
        ("Sensitivity Table: Discount Rate Sensitivity Range", "", "%")
    ]
    
    for i, (label, value, format) in enumerate(data_structure, 1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'C{i}'] = format
    
    title_font = Font(name='Arial', size=12, bold=True)
    for col in ['A', 'B', 'C']:
        ws[f'{col}1'].font = title_font
    
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    
    yes_no_validation = DataValidation(type="list", formula1='"Yes,No"', allow_blank=False)
    ws.add_data_validation(yes_no_validation)
    yes_no_validation.add(ws['B7'])
    yes_no_validation.add(ws['B9'])
    
    ws['B10'].number_format = '0%'
    ws['B11'].number_format = '0.00%'
    
    for row in [12, 13, 14]:
        ws[f'B{row}'].number_format = '_-* #,##0.00\\ _€_-;\\-* #,##0.00\\ _€_-;_-* "-"??\\ _€_-;_-@_-'
    
    curve_validation = DataValidation(type="list", 
                                    formula1='"Linear Curve,Exponential Curve,S-Curve"', 
                                    allow_blank=False)
    ws.add_data_validation(curve_validation)
    curve_validation.add(ws['B15'])
    
    ws['B16'].number_format = '0%'
    
    for row in [17, 18, 19]:
        ws[f'B{row}'].number_format = '0.00%'
    
    date_format = 'm/d/yy'
    ws['B3'].number_format = date_format
    ws['B4'].number_format = date_format
    
    for row in range(1, 22):
        for col in ['A', 'B', 'C']:
            cell = ws[f'{col}{row}']
            cell.alignment = Alignment(vertical='center', wrap_text=True)
    
    for row in range(1, 22):
        ws[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        ws[f'B{row}'].alignment = Alignment(horizontal='center', vertical='center')
        ws[f'C{row}'].alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for row in range(1, 22):
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].border = thin_border
    
    return wb

def load_loans_dataframe_from_stream(file_stream, sheet_name='Loans'):
    """Load loans dataframe from BytesIO stream instead of file path"""
    try:
        file_stream.seek(0)
        df = pd.read_excel(file_stream, sheet_name=sheet_name, header=0)
        return df
    except Exception as e:
        print(f"Error loading Loans DataFrame: {str(e)}")
        return None

def find_data_tape_file(downloaded_files):
    """Find the data tape file from downloaded files"""
    for file_stream, filename in downloaded_files:
        if any(keyword in filename.lower() for keyword in ['data', 'datatape', 'simple', 'loan']):
            return file_stream, filename
    
    for file_stream, filename in downloaded_files:
        if filename.lower().endswith(('.xlsx', '.xls')):
            return file_stream, filename
    
    return None, None

def get_valuation_and_max_maturity_dates(loans_df):
    valuation_date = pd.to_datetime('2020-09-05')
    
    max_maturity_date = None
    if loans_df is not None and 'Maturity Date' in loans_df.columns:
        maturity_dates = pd.to_datetime(loans_df['Maturity Date'], errors='coerce')
        max_maturity_date = maturity_dates.max()
    return valuation_date, max_maturity_date

def generate_month_end_dates_list(valuation_date, max_maturity_date):
    dates = []
    dates.append(valuation_date)
    
    current = valuation_date.replace(day=1)
    current = current + pd.DateOffset(months=1) - pd.DateOffset(days=1)
    
    if current <= valuation_date:
        current = current + pd.DateOffset(months=1)
        current = current.replace(day=1) + pd.DateOffset(months=1) - pd.DateOffset(days=1)
    
    while current.year < max_maturity_date.year or (current.year == max_maturity_date.year and current.month <= max_maturity_date.month):
        dates.append(current)
        current = current + pd.DateOffset(months=1)
        current = current.replace(day=1) + pd.DateOffset(months=1) - pd.DateOffset(days=1)
    
    return dates

def create_date_columns_with_percentage_format(worksheet, date_list, start_col=2, start_row=1):
    for i, date in enumerate(date_list):
        col_num = start_col + i
        cell = worksheet.cell(row=start_row, column=col_num, value=date.strftime('%m/%d/%Y'))
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        col_letter = openpyxl.utils.get_column_letter(col_num)
        worksheet.column_dimensions[col_letter].width = 12
    
    return len(date_list)

def analyze_loan_types(loans_df):
    if loans_df is None:
        return []
    
    loan_type_column = None
    possible_columns = ['Type of Loan', 'Loan Type', 'type of loan', 'loan type']
    
    for col_name in possible_columns:
        if col_name in loans_df.columns:
            loan_type_column = col_name
            break
    
    if loan_type_column:
        loan_types = loans_df[loan_type_column].dropna().unique()
        loan_types = sorted([str(lt) for lt in loan_types])
        return loan_types
    else:
        return []

def create_assumption_loans_sheet(workbook, loans_df, date_list):
    ws_loans = workbook.create_sheet(title="Assumption_Loans")
    
    loan_types = analyze_loan_types(loans_df)
    
    if not loan_types:
        loan_types = [
            "Medium / Long Term Loan", "RE Leasing", "Overdraft", "Syndicated Loan",
            "Called Bank Guarantee", "Uncalled Bank Guarantee", "Factoring", 
            "Residential Mortgage", "Credit Card", "Corporate/ Development Loan",
            "Current Account", "Non RE Leasing", "Discounted Bill/ Note", 
            "Consumer Loan", "Other", "Trade Finance", "Restructured Loan"
        ]
    
    ws_loans['A1'] = "Cost of Risk - Loan with Guarantee"
    ws_loans['A1'].font = Font(name='Arial', size=12, bold=True)
    
    create_date_columns_with_percentage_format(ws_loans, date_list, start_col=2, start_row=1)
    
    for i, loan_type in enumerate(loan_types, 2):
        ws_loans.cell(row=i, column=1, value=loan_type)
        ws_loans.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        for col_idx in range(2, len(date_list) + 2):
            cell = ws_loans.cell(row=i, column=col_idx, value="")
            cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    first_table_rows = len(loan_types) + 1
    first_table_cols = len(date_list) + 1
    
    for row in range(1, first_table_rows + 1):
        for col in range(1, first_table_cols + 1):
            ws_loans.cell(row=row, column=col).border = thin_border
    
    second_table_start_row = first_table_rows + 4
    
    ws_loans.cell(row=second_table_start_row, column=1, value="Prepayment Risk - Loan with Guarantee")
    ws_loans.cell(row=second_table_start_row, column=1).font = Font(name='Arial', size=12, bold=True)
    
    for i, date in enumerate(date_list):
        col_num = i + 2
        cell = ws_loans.cell(row=second_table_start_row, column=col_num, value=date.strftime('%m/%d/%Y'))
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for i, loan_type in enumerate(loan_types, second_table_start_row + 1):
        ws_loans.cell(row=i, column=1, value=loan_type)
        ws_loans.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        for col_idx in range(2, len(date_list) + 2):
            cell = ws_loans.cell(row=i, column=col_idx, value="")
            cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    second_table_rows = len(loan_types) + 1
    
    for row in range(second_table_start_row, second_table_start_row + second_table_rows):
        for col in range(1, first_table_cols + 1):
            ws_loans.cell(row=row, column=col).border = thin_border
    
    third_table_start_row = second_table_start_row + second_table_rows + 4
    
    third_table_headers = ["Types of Loans", "Discount Rate", "Non-interest fees (over undrawn commitment)", "Non-interest fees (over outstanding balance)", "Servicing Fee"]
    
    for col, header in enumerate(third_table_headers, 1):
        cell = ws_loans.cell(row=third_table_start_row, column=col, value=header)
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for i, loan_type in enumerate(loan_types, third_table_start_row + 1):
        ws_loans.cell(row=i, column=1, value=loan_type)
        ws_loans.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        for col_idx in range(2, 6):
            cell = ws_loans.cell(row=i, column=col_idx, value="")
            cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    third_table_rows = len(loan_types) + 1
    third_table_cols = 5
    
    for row in range(third_table_start_row, third_table_start_row + third_table_rows):
        for col in range(1, third_table_cols + 1):
            ws_loans.cell(row=row, column=col).border = thin_border
    
    fourth_table_start_row = third_table_start_row + third_table_rows + 4
    
    fourth_table_headers = [
        "Types of Loans", 
        "30+ years", 
        "20 year < x <= 30 years", 
        "15 year < x <= 20 years",
        "10 year < x <= 15 years",
        "5 year < x <= 10 years",
        "3 year < x <= 5 years", 
        "1 year < x <= 3 years",
        "180 < x <=360 days",
        "90 days < x <=180 days",
        "30 days < x <=90 days",
        "< 30 days"
    ]
    
    day_values = [
        "",
        "100000000000000000+",
        "10800",
        "7200",
        "5400",
        "3600",
        "1800",
        "1080",
        "360",
        "180",
        "90",
        "30"
    ]
    
    ws_loans.cell(row=fourth_table_start_row, column=1, value="Recovery Rate per Vintage")
    ws_loans.cell(row=fourth_table_start_row, column=1).font = Font(name='Arial', size=12, bold=True)
    ws_loans.cell(row=fourth_table_start_row, column=1).alignment = Alignment(horizontal='center', vertical='center')
    
    ws_loans.merge_cells(start_row=fourth_table_start_row, start_column=1, 
                        end_row=fourth_table_start_row + 1, end_column=1)
    
    for col, header in enumerate(fourth_table_headers[1:], 2):
        cell = ws_loans.cell(row=fourth_table_start_row, column=col, value=header)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        col_letter = openpyxl.utils.get_column_letter(col)
        ws_loans.column_dimensions[col_letter].width = 15
    
    for col, day_value in enumerate(day_values[1:], 2):
        cell = ws_loans.cell(row=fourth_table_start_row + 1, column=col, value=day_value)
        cell.font = Font(name='Arial', size=10, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    data_start_row = fourth_table_start_row + 2
    
    for i, loan_type in enumerate(loan_types, data_start_row):
        ws_loans.cell(row=i, column=1, value=loan_type)
        ws_loans.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        for col_idx in range(2, len(fourth_table_headers) + 1):
            cell = ws_loans.cell(row=i, column=col_idx, value="")
            cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    fourth_table_rows = len(loan_types) + 2
    fourth_table_cols = len(fourth_table_headers)
    
    for row in range(fourth_table_start_row, fourth_table_start_row + fourth_table_rows):
        for col in range(1, fourth_table_cols + 1):
            ws_loans.cell(row=row, column=col).border = thin_border
    
    ws_loans.column_dimensions['A'].width = 35
    
    for col_letter in ['B', 'C', 'D', 'E']:
        ws_loans.column_dimensions[col_letter].width = 18
    
    return len(loan_types), fourth_table_start_row + fourth_table_rows

def analyze_index_column_from_df(loans_df):
    try:
        if loans_df is None:
            return None
        
        valuation_date, max_maturity_date = get_valuation_and_max_maturity_dates(loans_df)
        
        if 'Index' in loans_df.columns:
            index_column = loans_df['Index']
            non_null_values = index_column.dropna()
            value_counts = non_null_values.value_counts().sort_values(ascending=False)
                
            return {
                'total_rows': len(loans_df),
                'total_index_values': len(non_null_values),
                'null_values': index_column.isnull().sum(),
                'unique_count': len(value_counts),
                'value_counts': value_counts.to_dict(),
                'value_counts_series': value_counts,
                'valuation_date': valuation_date,
                'max_maturity_date': max_maturity_date,
                'loans_df': loans_df
            }
        else:
            return None
            
    except Exception as e:
        print(f"Index analysis error: {str(e)}")
        return None

def create_index_analysis_sheet(workbook, analysis_result):
    if not analysis_result:
        return
    
    ws_analysis = workbook.create_sheet(title="Index_Analysis")
    
    ws_analysis['A1'] = "Index Type"
    ws_analysis['A1'].font = Font(name='Arial', size=12, bold=True)
    
    valuation_date = analysis_result['valuation_date']
    max_maturity_date = analysis_result['max_maturity_date']
    month_end_dates = generate_month_end_dates_list(valuation_date, max_maturity_date)
    
    total_columns = len(month_end_dates) + 10
    
    create_date_columns_with_percentage_format(ws_analysis, month_end_dates, start_col=2, start_row=1)
    
    value_counts = analysis_result['value_counts_series']
    
    for i, (index_type, count) in enumerate(value_counts.items(), 2):
        ws_analysis.cell(row=i, column=1, value=index_type)
        ws_analysis.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
    
    ws_analysis.column_dimensions['A'].width = 35
    
    for i in range(len(month_end_dates), len(month_end_dates) + 10):
        col_letter = openpyxl.utils.get_column_letter(i + 2)
        ws_analysis.column_dimensions[col_letter].width = 12
        
        for row in range(2, len(value_counts) + 2):
            cell = ws_analysis.cell(row=row, column=i + 2)
            cell.number_format = '0.00%'
            cell.alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    total_rows = len(value_counts) + 1
    
    for row in range(1, total_rows + 1):
        for col_idx in range(1, total_columns + 1):
            ws_analysis.cell(row=row, column=col_idx).border = thin_border

def analyze_currencies_from_df(loans_df, filename=""):
    try:
        filename = filename.lower()
        loans_currencies = set()
        
        if loans_df is not None:
            loans_df.columns = loans_df.columns.str.strip()
            
            if 'Currency' in loans_df.columns:
                loans_currencies = set(loans_df['Currency'].dropna().unique())
            else:
                currency_like_columns = [col for col in loans_df.columns if 'currency' in str(col).lower() or 'curr' in str(col).lower()]
                if currency_like_columns:
                    currency_col = currency_like_columns[0]
                    loans_currencies = set(loans_df[currency_col].dropna().unique())
        
        all_currencies = sorted(list(loans_currencies)) if loans_currencies else ['EUR', 'USD', 'GBP']
        return all_currencies
        
    except Exception as e:
        print(f"Currency analysis error: {str(e)}")
        return ['EUR', 'USD', 'GBP']

def analyze_guarantee_types_from_df(loans_df, filename=""):
    try:
        filename = filename.lower()
        guarantee_types = []
        
        if loans_df is not None:
            loans_df.columns = [str(col).strip() if pd.notna(col) else col for col in loans_df.columns]

            possible_columns = ['Type of Guarantees', 'Type of Guarantee', 'Guarantees Type', 'Guarantee Type']
            
            for col_name in possible_columns:
                if col_name in loans_df.columns:
                    guarantee_types = loans_df[col_name].dropna().unique()
                    guarantee_types = sorted([str(gt).strip() for gt in guarantee_types])
                    break
            
            if not guarantee_types:
                guarantee_cols = [col for col in loans_df.columns if pd.notna(col) and 'guarantee' in str(col).lower()]
                if guarantee_cols:
                    col_name = guarantee_cols[-1]
                    guarantee_types = loans_df[col_name].dropna().unique()
                    guarantee_types = sorted([str(gt).strip() for gt in guarantee_types])
        
        return guarantee_types
        
    except Exception as e:
        print(f"Guarantee types analysis error: {str(e)}")
        return []

def generate_currency_pairs(currencies):
    currency_pairs = []
    
    for quote_currency in currencies:
        for base_currency in currencies:
            if quote_currency != base_currency:
                currency_pairs.append((quote_currency, base_currency))
    
    return currency_pairs

def add_guarantee_recovery_table_to_loans_sheet(ws_loans, guarantee_types, start_row):
    table_start_row = start_row + 3

    headers = ["Types of Guarantee", "Recovery Rate", "Accepted Format"]
    for col, header in enumerate(headers, 1):
        cell = ws_loans.cell(row=table_start_row, column=col, value=header)
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for i, guarantee_type in enumerate(guarantee_types, table_start_row + 1):
        ws_loans.cell(row=i, column=1, value=guarantee_type)
        ws_loans.cell(row=i, column=1).alignment = Alignment(horizontal='left', vertical='center')
        
        rate_cell = ws_loans.cell(row=i, column=2, value="")
        rate_cell.number_format = '0.00%'
        rate_cell.alignment = Alignment(horizontal='center', vertical='center')

        ws_loans.cell(row=i, column=3, value="%")
        ws_loans.cell(row=i, column=3).alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    table_rows = len(guarantee_types) + 1
    for row in range(table_start_row, table_start_row + table_rows):
        for col in range(1, 4):
            ws_loans.cell(row=row, column=col).border = thin_border
    
    return len(guarantee_types)

def create_assumption_currency_sheet(workbook, currencies, loans_df=None):
    ws_currency = workbook.create_sheet(title="Assumption_Currency")
    
    currency_pairs = generate_currency_pairs(currencies)
    
    headers = ["Quote Currency", "Base Currency", "Exchange Rate at Valuation Date"]
    for col, header in enumerate(headers, 1):
        cell = ws_currency.cell(row=1, column=col, value=header)
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for row, (quote_currency, base_currency) in enumerate(currency_pairs, 2):
        ws_currency.cell(row=row, column=1, value=quote_currency)
        ws_currency.cell(row=row, column=2, value=base_currency)
        
        rate_cell = ws_currency.cell(row=row, column=3, value="")
        rate_cell.number_format = '0.0000'
        rate_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    first_table_rows = len(currency_pairs) + 1
    for row in range(1, first_table_rows + 1):
        for col in range(1, 4):
            ws_currency.cell(row=row, column=col).border = thin_border
    
    second_table_start_row = first_table_rows + 4
    
    loans_currencies = []
    if loans_df is not None:
        if 'Currency' in loans_df.columns:
            loans_currencies = sorted(loans_df['Currency'].dropna().unique().tolist())
        else:
            currency_like_columns = [col for col in loans_df.columns if 'currency' in str(col).lower() or 'curr' in str(col).lower()]
            if currency_like_columns:
                currency_col = currency_like_columns[0]
                loans_currencies = sorted(loans_df[currency_col].dropna().unique().tolist())
    
    if not loans_currencies:
        loans_currencies = currencies[:8] if len(currencies) >= 8 else currencies
    
    second_table_headers = ["Local Currency (Performing Loans only)", "Corporate Tax"]
    
    for col, header in enumerate(second_table_headers, 1):
        cell = ws_currency.cell(row=second_table_start_row, column=col, value=header)
        cell.font = Font(name='Arial', size=12, bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    for i, currency in enumerate(loans_currencies, second_table_start_row + 1):
        ws_currency.cell(row=i, column=1, value=currency)
        ws_currency.cell(row=i, column=1).alignment = Alignment(horizontal='center', vertical='center')
        
        cell = ws_currency.cell(row=i, column=2, value="")
        cell.number_format = '0.00%'
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    second_table_rows = len(loans_currencies) + 1
    
    for row in range(second_table_start_row, second_table_start_row + second_table_rows):
        for col in range(1, 3):
            ws_currency.cell(row=row, column=col).border = thin_border
    
    ws_currency.column_dimensions['A'].width = 30
    ws_currency.column_dimensions['B'].width = 18
    ws_currency.column_dimensions['C'].width = 25
    
    return len(currency_pairs)

def upload_excel_to_sharepoint(workbook, output_folder_path):
    """Upload Excel workbook to SharePoint Phase 2 folder"""
    try:
        # Convert workbook to BytesIO
        output_buffer = BytesIO()
        workbook.save(output_buffer)
        output_buffer.seek(0)
        
        # Get SharePoint connection
        env_vars = load_env_vars()
        access_token = get_access_token(env_vars)
        site_id = get_site_id(access_token, env_vars['site_name'])
        
        # Create filename with timestamp
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Assumption_Summary_Template_{timestamp}.xlsx"
        
        # Build output path for Phase 2 - now using full path
        item_path = f"{output_folder_path}/{filename}"
        
        # Upload
        size = len(output_buffer.getvalue())
        
        if size <= 4 * 1024 * 1024:  # <= 4MB -> simple upload
            url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{item_path}:/content"
            headers = {
                "Authorization": f"Bearer {access_token}",
                "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
            import requests
            resp = requests.put(url, headers=headers, data=output_buffer.getvalue())
            resp.raise_for_status()
            return resp.json()
        
        # For larger files, use chunked upload
        session_url = f"https://graph.microsoft.com/v1.0/sites/{site_id}/drive/root:/{item_path}:/createUploadSession"
        session_headers = {"Authorization": f"Bearer {access_token}"}
        session_body = {
            "item": {
                "@microsoft.graph.conflictBehavior": "replace",
                "name": filename,
            }
        }
        session_resp = requests.post(session_url, headers=session_headers, json=session_body)
        session_resp.raise_for_status()
        upload_url = session_resp.json().get("uploadUrl")

        chunk_size = 5 * 1024 * 1024  # 5 MB
        start = 0
        data = output_buffer.getvalue()
        last_resp = None

        while start < size:
            end = min(start + chunk_size, size) - 1
            chunk = data[start : end + 1]
            headers = {
                "Content-Length": str(end - start + 1),
                "Content-Range": f"bytes {start}-{end}/{size}",
            }
            put_resp = requests.put(upload_url, headers=headers, data=chunk)
            put_resp.raise_for_status()
            last_resp = put_resp
            start = end + 1

        return last_resp.json()
        
    except Exception as e:
        print(f"Error uploading Excel to SharePoint: {str(e)}")
        return None

def process_sharepoint_assumption_template():
    """
    Main function to download from Phase 1, create assumption template, and upload to Phase 2
    """
    print("Starting SharePoint integrated assumption template generation...")
    
    try:
        # 1. Load environment variables and authenticate
        env_vars = load_env_vars()
        access_token = get_access_token(env_vars)
        site_id = get_site_id(access_token, env_vars["site_name"])
        
        # 2. Define folder paths - using full path like your friend's example
        input_folder_path = f"{env_vars['base_path']}/Phase 1"
        output_folder_path = f"{env_vars['base_path']}/Phase 2"
        
        # 3. Download files from SharePoint Phase 1
        print(f"Downloading files from: {input_folder_path}")
        downloaded_files = download_file(access_token, site_id, input_folder_path)
        
        # 4. Find the data tape file
        data_tape_stream, data_tape_filename = find_data_tape_file(downloaded_files)
        
        if not data_tape_stream:
            print("No data tape file found in Phase 1. Creating basic template only...")
            wb = create_assumption_summary_excel()
            
        else:
            print(f"Found data tape file: {data_tape_filename}")
            
            # 5. Load and analyze the data
            loans_df = load_loans_dataframe_from_stream(data_tape_stream)
            
            if loans_df is None:
                print("Could not load loans data. Creating basic template only...")
                wb = create_assumption_summary_excel()
            else:
                print("Analyzing loans data...")
                
                # 6. Perform all analyses
                valuation_date, max_maturity_date = get_valuation_and_max_maturity_dates(loans_df)
                date_list = generate_month_end_dates_list(valuation_date, max_maturity_date)
                
                analysis_result = analyze_index_column_from_df(loans_df)
                currencies = analyze_currencies_from_df(loans_df, data_tape_filename)
                guarantee_types = analyze_guarantee_types_from_df(loans_df, data_tape_filename)
                
                # 7. Create Excel workbook with all sheets
                wb = create_assumption_summary_excel()
                print(f"Base workbook created with sheets: {[sheet.title for sheet in wb.worksheets]}")
                
                # Add Assumption_Loans sheet
                if loans_df is not None and date_list:
                    loan_count, last_table_end = create_assumption_loans_sheet(wb, loans_df, date_list)
                    print(f"Created Assumption_Loans sheet with {loan_count} loan types")
                    print(f"Current sheets after Assumption_Loans: {[sheet.title for sheet in wb.worksheets]}")
                    
                    # Add guarantee recovery table if guarantee types found
                    if guarantee_types:
                        ws_loans = wb["Assumption_Loans"]
                        add_guarantee_recovery_table_to_loans_sheet(ws_loans, guarantee_types, last_table_end)
                        print(f"Added guarantee recovery table with {len(guarantee_types)} guarantee types")
                
                # Add Index_Analysis sheet
                if analysis_result:
                    create_index_analysis_sheet(wb, analysis_result)
                    print(f"Created Index_Analysis sheet with {analysis_result['unique_count']} index types")
                    print(f"Current sheets after Index_Analysis: {[sheet.title for sheet in wb.worksheets]}")
                
                # Add Assumption_Currency sheet
                if currencies:
                    create_assumption_currency_sheet(wb, currencies, loans_df)
                    print(f"Created Assumption_Currency sheet with {len(currencies)} currencies")
                    print(f"Final sheets before upload: {[sheet.title for sheet in wb.worksheets]}")
                
                print("Assumption template created with all analyzed data")
        
        # 8. Upload Excel workbook to SharePoint Phase 2
        print(f"Uploading assumption template to: {output_folder_path}")
        result = upload_excel_to_sharepoint(wb, output_folder_path)
        
        if result:
            print("SUCCESS: Assumption template successfully uploaded to SharePoint Phase 2!")
            print(f"File location: {output_folder_path}")
            return result
        else:
            print("ERROR: Failed to upload assumption template")
            return None
        
    except Exception as e:
        print(f"Error in SharePoint assumption template process: {str(e)}")
        return None

if __name__ == "__main__":
    print("Starting SharePoint Assumption Template Generator")
    print("=" * 60)
    print("Configuration:")
    print("- Input folder: base_path/Phase 1")
    print("- Output folder: base_path/Phase 2")
    print("- Base path from .env: LGG Guides/Teneo-Testing")
    print("=" * 60)
    
    # Process with SharePoint Phase 1 -> Phase 2 workflow
    result = process_sharepoint_assumption_template()
    
    if result:
        print("\n" + "=" * 60)
        print("PROCESS COMPLETED SUCCESSFULLY!")
        print("=" * 60)
        print("Summary of actions:")
        print("✓ Downloaded Excel files from SharePoint Phase 1")
        print("✓ Analyzed loan data from data tape file")
        print("✓ Created comprehensive assumption template")
        print("✓ Generated all required sheets:")
        print("  - Assumption_Summary")
        print("  - Assumption_Loans (with loan types analysis)")
        print("  - Index_Analysis (if index data found)")
        print("  - Assumption_Currency (with currency pairs)")
        print("  - Guarantee Recovery Tables (if guarantee data found)")
        print("✓ Uploaded final template to SharePoint Phase 2")
        print("=" * 60)
    else:
        print("\n" + "=" * 60)
        print("PROCESS FAILED!")
        print("=" * 60)
        print("Please check the error messages above and ensure:")
        print("- SharePoint credentials are correct in .env file")
        print("- Phase 1 folder exists and contains Excel files")
        print("- Phase 2 folder exists and is writable")
        print("- Network connection is stable")
        print("=" * 60)